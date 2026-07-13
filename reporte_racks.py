import os
import json
from datetime import datetime
from decimal import Decimal

import boto3
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from dotenv import load_dotenv


load_dotenv()


def convertir_dynamodb(item):
    """Convierte formato de tipos de DynamoDB al formato normal JSON."""
    if item is None:
        return None
    
    nuevo = {}
    for clave, valor in item.items():
        if isinstance(valor, dict):
            if "N" in valor:
                val = Decimal(valor["N"])
                nuevo[clave] = int(val) if val % 1 == 0 else float(val)
            elif "S" in valor:
                nuevo[clave] = valor["S"]
            elif "L" in valor:
                nuevo[clave] = [convertir_dynamodb({"item": v})["item"] for v in valor["L"]]
            elif "M" in valor:
                nuevo[clave] = convertir_dynamodb(valor["M"])
            elif "BOOL" in valor:
                nuevo[clave] = valor["BOOL"]
            elif "NULL" in valor:
                nuevo[clave] = None
            else:
                nuevo[clave] = valor
        else:
            nuevo[clave] = valor
    return nuevo


def obtener_datos_buffer(tabla_nombre, region):
    """Obtiene todos los datos de la tabla Buffer."""
    try:
        session = boto3.Session(region_name=region)
        dynamodb_client = session.client("dynamodb", region_name=region)
        
        respuesta = dynamodb_client.scan(TableName=tabla_nombre)
        items = respuesta.get("Items", [])
        
        return [convertir_dynamodb(item) for item in items]
    
    except Exception as e:
        print(f"Error leyendo tabla {tabla_nombre}: {e}")
        return []


def obtener_receta(nombre_receta, tabla_nombre, region):
    """Obtiene datos de una receta específica por nombre."""
    try:
        session = boto3.Session(region_name=region)
        dynamodb_client = session.client("dynamodb", region_name=region)
        
        respuesta = dynamodb_client.query(
            TableName=tabla_nombre,
            KeyConditionExpression="nombre_receta = :nombre",
            ExpressionAttributeValues={
                ":nombre": {"S": nombre_receta}
            }
        )
        
        items = respuesta.get("Items", [])
        if items:
            return convertir_dynamodb(items[0])
        return None
    
    except Exception as e:
        print(f"Error obteniendo receta {nombre_receta}: {e}")
        return None


def obtener_todas_recetas(tabla_nombre, region):
    """Obtiene todas las recetas de la tabla."""
    try:
        session = boto3.Session(region_name=region)
        dynamodb_client = session.client("dynamodb", region_name=region)
        
        respuesta = dynamodb_client.scan(TableName=tabla_nombre)
        items = respuesta.get("Items", [])
        
        recetas = {}
        for item in items:
            receta_convertida = convertir_dynamodb(item)
            if "nombre_receta" in receta_convertida:
                recetas[receta_convertida["nombre_receta"]] = receta_convertida
        
        return recetas
    
    except Exception as e:
        print(f"Error leyendo tabla de recetas {tabla_nombre}: {e}")
        return {}


def parsear_nivel_json(nivel_str):
    """
    Parsea un string JSON de nivel y retorna los valores.
    Estructura: [cancelaciones, finalizado, seleccionado, tiempo_segundos]
    """
    try:
        if not nivel_str:
            return {"cancelaciones": 0, "finalizado": 0, "seleccionado": 0, "tiempo_segundos": 0}
        
        # Si es string, parsear JSON
        if isinstance(nivel_str, str):
            datos = json.loads(nivel_str)
        else:
            datos = nivel_str
        
        # Parsear el formato DynamoDB dentro del JSON
        cancelaciones = datos[0].get("N", 0) if isinstance(datos[0], dict) else datos[0]
        finalizado = datos[1].get("N", 0) if isinstance(datos[1], dict) else datos[1]
        seleccionado = datos[2].get("N", 0) if isinstance(datos[2], dict) else datos[2]
        tiempo_segundos = datos[3].get("N", 0) if isinstance(datos[3], dict) else datos[3]
        
        return {
            "cancelaciones": int(cancelaciones),
            "finalizado": int(finalizado),
            "seleccionado": int(seleccionado),
            "tiempo_segundos": int(tiempo_segundos)
        }
    except:
        return {"cancelaciones": 0, "finalizado": 0, "seleccionado": 0, "tiempo_segundos": 0}


def calcular_tiempo_diferencia(fecha_inicio, fecha_fin):
    """Calcula la diferencia entre dos fechas y retorna en formato HH:MM:SS"""
    if not fecha_inicio or not fecha_fin:
        return "00:00:00"
    
    try:
        # Si son strings, convertir a datetime
        if isinstance(fecha_inicio, str):
            fecha_inicio = datetime.strptime(fecha_inicio, "%d-%m-%Y %H:%M:%S")
        if isinstance(fecha_fin, str):
            fecha_fin = datetime.strptime(fecha_fin, "%d-%m-%Y %H:%M:%S")
        
        diferencia = fecha_fin - fecha_inicio
        segundos_totales = int(diferencia.total_seconds())
        
        if segundos_totales < 0:
            return "00:00:00"
        
        horas = segundos_totales // 3600
        minutos = (segundos_totales % 3600) // 60
        segundos = segundos_totales % 60
        
        return f"{horas:02d}:{minutos:02d}:{segundos:02d}"
    except:
        return "00:00:00"


def procesar_datos_racks(buffer_data, recetas):
    """Procesa los datos del buffer para generar el reporte."""
    racks_data = []
    niveles_data = []
    
    for buffer_item in buffer_data:
        fecha_inicio = buffer_item.get("fecha_inicio", "")
        fecha_fin = buffer_item.get("fecha_fin", "")
        equipo = buffer_item.get("equipoSeleccionado", "")
        rack = buffer_item.get("rackBuffer1", "")
        receta_nombre = buffer_item.get("recetaBuffer1", "")
        
        receta = recetas.get(receta_nombre, {})
        peso_producto = receta.get("peso_producto", 0) or 0
        productos_nivel = receta.get("productos_nivel", 0) or 0
        
        try:
            peso_producto = float(peso_producto) if peso_producto else 0
            productos_nivel = int(productos_nivel) if productos_nivel else 0
        except:
            peso_producto = 0
            productos_nivel = 0
        
        niveles_finalizados = 0
        tiempo_total_segundos = 0
        
        for num_nivel in range(1, 14):
            nivel_key = f"nivel{num_nivel}"
            nivel_data = buffer_item.get(nivel_key)
            
            if nivel_data:
                nivel_info = parsear_nivel_json(nivel_data)
                
                if nivel_info["finalizado"] == 1:
                    niveles_finalizados += 1
                
                niveles_data.append({
                    "rack": rack,
                    "numero_nivel": num_nivel,
                    "cancelaciones": nivel_info["cancelaciones"],
                    "finalizado": "SI" if nivel_info["finalizado"] == 1 else "NO",
                    "seleccionado": "SI" if nivel_info["seleccionado"] == 1 else "NO",
                    "tiempo_segundos": nivel_info["tiempo_segundos"]
                })
        
        kilos_procesados = niveles_finalizados * peso_producto * productos_nivel
        
        tiempo_neto = calcular_tiempo_diferencia(fecha_inicio, fecha_fin)
        
        fecha_inicio_str = fecha_inicio if isinstance(fecha_inicio, str) else ""
        fecha_fin_str = fecha_fin if isinstance(fecha_fin, str) else ""
        
        racks_data.append({
            "fecha_inicio": fecha_inicio_str,
            "fecha_fin": fecha_fin_str,
            "tiempo_neto": tiempo_neto,
            "equipo": equipo,
            "rack": rack,
            "receta": receta_nombre,
            "niveles_finalizados": niveles_finalizados,
            "kilos_procesados": f"{kilos_procesados:.2f}"
        })
    
    return racks_data, niveles_data


def export_racks_to_excel(file_path, tabla_buffer, tabla_receta, region):
    """
    Genera un Excel con el reporte de racks.
    Args:
        file_path: Ruta del archivo Excel a generar
        tabla_buffer: Nombre de la tabla Buffer en DynamoDB
        tabla_receta: Nombre de la tabla Receta en DynamoDB
        region: Región de AWS
    Returns:
        True si se generó correctamente, False si no hay datos
    """
    print("Exportando datos de racks a Excel...")
    
    buffer_data = obtener_datos_buffer(tabla_buffer, region)
    recetas = obtener_todas_recetas(tabla_receta, region)
    
    if not buffer_data:
        print("No se encontraron datos en la tabla Buffer")
        return False
    
    if not recetas:
        print("No se encontraron recetas")
        return False
    
    try:
        racks_data, niveles_data = procesar_datos_racks(buffer_data, recetas)
        
        wb = Workbook()
        ws_racks = wb.active
        ws_racks.title = "Racks"
        
        # Ruta absoluta del logo
        logo_dir = os.path.dirname(os.path.abspath(__file__))
        logo_path = os.path.join(logo_dir, "static", "cremonarecort.png")
        
        if os.path.exists(logo_path):
            try:
                img = XLImage(logo_path)
                img.height = 31.5
                img.width = 126
                ws_racks.add_image(img, "D3")
            except Exception as logo_error:
                print(f"Advertencia: No se pudo agregar el logo: {logo_error}")
        
        ws_racks.merge_cells("A1:H1")
        ws_racks["A1"] = "REPORTE DE RACKS"
        ws_racks["A1"].font = Font(size=16, bold=True, color="FFFFFF")
        ws_racks["A1"].alignment = Alignment(horizontal="center")
        ws_racks["A1"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        ws_racks["A3"] = ""
        ws_racks["A4"] = ""
        
        ws_racks.append([])
        
        # Headers para Racks
        headers_racks = [
            "Fecha de inicio de la tarea\n[YYYY-MM-DD HH:MM:SS]",
            "Fecha de finalización de la tarea\n[YYYY-MM-DD HH:MM:SS]",
            "Tiempo neto [HH:MM:SS]",
            "Equipo",
            "Rack",
            "Receta",
            "Niveles finalizados",
            "Kilos procesados [KG]"
        ]
        
        ws_racks.append(headers_racks)
        first_table_first_row = ws_racks.max_row
        
        # Formatear headers
        for col in range(1, len(headers_racks) + 1):
            cell = ws_racks.cell(row=first_table_first_row, column=col)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        ws_racks.row_dimensions[first_table_first_row].height = 45
        
        # Agregar datos de racks
        for rack_data in racks_data:
            ws_racks.append([
                rack_data["fecha_inicio"],
                rack_data["fecha_fin"],
                rack_data["tiempo_neto"],
                rack_data["equipo"],
                rack_data["rack"],
                rack_data["receta"],
                rack_data["niveles_finalizados"],
                rack_data["kilos_procesados"]
            ])
        
        first_table_last_row = ws_racks.max_row
        
        # Crear tabla para Racks
        first_table = Table(displayName="ReporteRacks", ref=f"A{first_table_first_row}:H{first_table_last_row}")
        first_style = TableStyleInfo(
            name="TableStyleMedium9", showFirstColumn=False,
            showLastColumn=False, showRowStripes=True, showColumnStripes=False
        )
        first_table.tableStyleInfo = first_style
        ws_racks.add_table(first_table)
        
        # Ajustar anchos de columna
        ws_racks.column_dimensions['A'].width = 32
        ws_racks.column_dimensions['B'].width = 32
        ws_racks.column_dimensions['C'].width = 18
        ws_racks.column_dimensions['D'].width = 15
        ws_racks.column_dimensions['E'].width = 12
        ws_racks.column_dimensions['F'].width = 20
        ws_racks.column_dimensions['G'].width = 18
        ws_racks.column_dimensions['H'].width = 20
        
        # ========== SEGUNDA HOJA: NIVELES RACKS ==========
        
        ws_niveles = wb.create_sheet("Niveles Racks")
        
        # Título
        ws_niveles.merge_cells("A1:E1")
        ws_niveles["A1"] = "REPORTE DE NIVELES RACKS"
        ws_niveles["A1"].font = Font(size=16, bold=True, color="FFFFFF")
        ws_niveles["A1"].alignment = Alignment(horizontal="center")
        ws_niveles["A1"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        ws_niveles["A3"] = ""
        ws_niveles["A4"] = ""
        
        ws_niveles.append([])
        
        # Headers para Niveles
        headers_niveles = [
            "Número de nivel",
            "Número de cancelaciones",
            "Finalizado [SI/NO]",
            "Seleccionado [SI/NO]",
            "TiempoNivel [seg]"
        ]
        
        ws_niveles.append(headers_niveles)
        second_table_first_row = ws_niveles.max_row
        
        # Formatear headers
        for col in range(1, len(headers_niveles) + 1):
            cell = ws_niveles.cell(row=second_table_first_row, column=col)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        ws_niveles.row_dimensions[second_table_first_row].height = 45
        
        # Agregar datos de niveles
        for nivel_data in niveles_data:
            ws_niveles.append([
                nivel_data["numero_nivel"],
                nivel_data["cancelaciones"],
                nivel_data["finalizado"],
                nivel_data["seleccionado"],
                nivel_data["tiempo_segundos"]
            ])
        
        second_table_last_row = ws_niveles.max_row
        
        # Crear tabla para Niveles
        second_table = Table(displayName="ReporteNivelesRacks", ref=f"A{second_table_first_row}:E{second_table_last_row}")
        second_style = TableStyleInfo(
            name="TableStyleMedium9", showFirstColumn=False,
            showLastColumn=False, showRowStripes=True, showColumnStripes=False
        )
        second_table.tableStyleInfo = second_style
        ws_niveles.add_table(second_table)
        
        # Ajustar anchos de columna
        ws_niveles.column_dimensions['A'].width = 18
        ws_niveles.column_dimensions['B'].width = 24
        ws_niveles.column_dimensions['C'].width = 20
        ws_niveles.column_dimensions['D'].width = 20
        ws_niveles.column_dimensions['E'].width = 18
        
        # Guardar archivo
        wb.save(file_path)
        print(f"Excel generado exitosamente en: {file_path}")
        return True
        
    except Exception as e:
        print(f"Error al generar el Excel: {str(e)}")
        import traceback
        traceback.print_exc()
        return False


if __name__ == "__main__":
    # Parámetros de configuración
    TABLA_BUFFER = os.getenv("TABLA_BUFFER", "Buffer")
    TABLA_RECETA = os.getenv("TABLA_RECETA", "Receta")
    REGION = os.getenv("AWS_REGION", "sa-east-1")
    
    # Generar archivo Excel
    output_path = "reporte_racks.xlsx"
    export_racks_to_excel(output_path, TABLA_BUFFER, TABLA_RECETA, REGION)
